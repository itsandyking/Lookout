"""Matrixify integration — import Shopify state, enrich images, export for upload.

Three main classes:
- MatrixifyImporter: Backfill variant images/positions from a Matrixify XLSX export
- ImageEnricher: Fill image gaps using catalog data (barcode + style-map matching)
- MatrixifyExporter: Export enrichments as Matrixify-compatible CSV for Shopify upload
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from pathlib import Path

from lookout.store import LookoutStore

# ---------------------------------------------------------------------------
# MatrixifyImporter
# ---------------------------------------------------------------------------


@dataclass
class MatrixifyImportResult:
    """Result of importing a Matrixify XLSX export."""

    variants_updated: int = 0
    variants_skipped: int = 0  # not found in DB
    images_set: int = 0
    positions_set: int = 0
    products_seen: int = 0


class MatrixifyImporter:
    """Import variant image assignments and positions from a Matrixify XLSX export.

    TODO: This class writes directly to the DB via the underlying TVR session.
    It is a compromise until TVR provides proper write APIs. Once TVR exposes
    update_variant() or similar, this should be rewritten to use LookoutStore
    exclusively.
    """

    def __init__(self, store: LookoutStore) -> None:
        self.store = store

    def import_file(self, path: str | Path) -> MatrixifyImportResult:
        """Read a Matrixify Products XLSX and update variant image_src/position."""
        import openpyxl

        # Access underlying TVR store session for DB writes
        from tvr.db.models import Variant

        path = Path(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Products"] if "Products" in wb.sheetnames else wb.active

        # Read headers from row 1
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(h) if h else "" for h in row]

        col = {h: i for i, h in enumerate(headers)}
        result = MatrixifyImportResult()
        product_ids_seen = set()

        with self.store._store.session() as session:
            for row in ws.iter_rows(min_row=2, values_only=True):
                values = list(row)
                if "Variant ID" not in col:
                    continue
                if not values or not values[col["Variant ID"]]:
                    continue

                variant_id = values[col["Variant ID"]]
                try:
                    variant_id = int(variant_id)
                except (ValueError, TypeError):
                    result.variants_skipped += 1
                    continue

                variant = session.query(Variant).filter(Variant.id == variant_id).first()

                if not variant:
                    result.variants_skipped += 1
                    continue

                updated = False

                # Image
                image_url = values[col.get("Variant Image", -1)] if "Variant Image" in col else None
                if image_url and str(image_url).strip():
                    variant.image_src = str(image_url).strip()
                    result.images_set += 1
                    updated = True

                # Position
                position = (
                    values[col.get("Variant Position", -1)] if "Variant Position" in col else None
                )
                if position is not None:
                    try:
                        variant.position = int(position)
                        result.positions_set += 1
                        updated = True
                    except (ValueError, TypeError):
                        pass

                if updated:
                    result.variants_updated += 1

                # Track products
                product_id = values[col.get("ID", 0)]
                if product_id:
                    product_ids_seen.add(product_id)

            session.commit()

        result.products_seen = len(product_ids_seen)
        wb.close()
        return result


# ---------------------------------------------------------------------------
# ImageEnricher
# ---------------------------------------------------------------------------


@dataclass
class ImageEnrichmentResult:
    """Result of the image enrichment pipeline."""

    propagated_from_shopify: int = 0  # Step A — same-color propagation
    matched_from_catalog: int = 0  # Step B — barcode lookup
    matched_from_style_map: int = 0  # Step C — style+color match
    still_missing: int = 0  # Could not find image
    products_processed: int = 0
    description_enrichments: list[dict] = field(default_factory=list)
    assignments: list[dict] = field(default_factory=list)
    # [{variant_id, product_id, product_handle, image_url, source, source_id}, ...]


class ImageEnricher:
    """Enrich variant images using existing Shopify data and catalog database.

    Three-step pipeline per color group:
    A. Propagate images from variants that already have them (same color)
    B. Look up variant barcodes in catalog_items for image_url
    C. Fall back to style-map + color-name matching against catalog
    """

    def __init__(self, store: LookoutStore) -> None:
        self.store = store

    def enrich(
        self,
        vendor: str | None = None,
        dry_run: bool = False,
    ) -> ImageEnrichmentResult:
        """Run the full enrichment pipeline.

        Args:
            vendor: Optional vendor filter
            dry_run: If True, don't write to DB, just collect assignments
        """
        result = ImageEnrichmentResult()

        products = self.store.list_products(vendor=vendor)

        for product in products:
            variants = self.store.get_variants(product["id"])
            if not variants:
                continue

            result.products_processed += 1
            self._enrich_product(product, variants, result, dry_run)

            # Description enrichment
            if not product["body_html"] or not product["body_html"].strip():
                desc = self.store.find_catalog_description(product["id"])
                if desc:
                    result.description_enrichments.append(
                        {
                            "product_id": product["id"],
                            "product_handle": product["handle"],
                            "body_html": desc,
                        }
                    )

        return result

    def _enrich_product(
        self,
        product: dict,
        variants: list[dict],
        result: ImageEnrichmentResult,
        dry_run: bool,
    ) -> None:
        """Enrich a single product's variants with images."""
        # Determine if this product uses Color as Option1
        has_color_option = any(
            v["option1_name"] and v["option1_name"].lower() == "color" for v in variants
        )

        if has_color_option:
            # Group variants by color (Option1 Value)
            color_groups: dict[str, list[dict]] = {}
            for v in variants:
                color = v["option1_value"] or "Unknown"
                color_groups.setdefault(color, []).append(v)

            for color, group in color_groups.items():
                self._enrich_color_group(product, group, color, result, dry_run)
        else:
            # No color grouping — treat each variant independently
            for v in variants:
                if v["image_src"]:
                    continue  # Already has image
                self._enrich_single_variant(product, v, result, dry_run)

    def _enrich_color_group(
        self,
        product: dict,
        variants: list[dict],
        color: str,
        result: ImageEnrichmentResult,
        dry_run: bool,
    ) -> None:
        """Enrich a color group of variants."""
        # Split into has-image vs needs-image
        with_image = [v for v in variants if v["image_src"]]
        without_image = [v for v in variants if not v["image_src"]]

        if not without_image:
            return  # All variants already have images

        # Step A: Same-color propagation
        if with_image:
            source_variant = with_image[0]
            image_url = source_variant["image_src"]
            for v in without_image:
                self._record_assignment(
                    v,
                    product,
                    image_url,
                    source="shopify_propagation",
                    source_id=source_variant["id"],
                    result=result,
                )
                result.propagated_from_shopify += 1
            return

        # Step B: Catalog barcode lookup
        image_url = None
        source_catalog_id = None
        for v in variants:
            barcode = v["barcode"]
            if not barcode:
                continue
            found_url = self.store.find_catalog_image(barcode)
            if found_url:
                image_url = found_url
                source_catalog_id = None  # LookoutStore doesn't expose catalog item IDs
                break

        if image_url:
            for v in without_image:
                self._record_assignment(
                    v,
                    product,
                    image_url,
                    source="catalog_barcode",
                    source_id=source_catalog_id,
                    result=result,
                )
                result.matched_from_catalog += 1
            return

        # Step C: Style-map + color match
        image_url = self.store.find_catalog_image_by_style(product["vendor"], "", color)
        # Note: find_catalog_image_by_style needs (vendor, style, color) but
        # we don't have the style code from the product dict. The store's
        # find_catalog_image_by_style looks up style via VendorStyleMap internally
        # only when called with the right style code. For now this may not match.
        # TODO: Add a find_catalog_image_for_product(product_id, color) method to LookoutStore.
        if image_url:
            for v in without_image:
                self._record_assignment(
                    v,
                    product,
                    image_url,
                    source="catalog_style_color",
                    source_id=None,
                    result=result,
                )
                result.matched_from_style_map += 1
            return

        # No image found
        result.still_missing += len(without_image)

    def _enrich_single_variant(
        self,
        product: dict,
        variant: dict,
        result: ImageEnrichmentResult,
        dry_run: bool,
    ) -> None:
        """Enrich a single variant (no color grouping)."""
        # Step B: Catalog barcode lookup
        barcode = variant["barcode"]
        if barcode:
            image_url = self.store.find_catalog_image(barcode)
            if image_url:
                self._record_assignment(
                    variant,
                    product,
                    image_url,
                    source="catalog_barcode",
                    source_id=None,
                    result=result,
                )
                result.matched_from_catalog += 1
                return

        # No image found
        result.still_missing += 1

    def _record_assignment(
        self,
        variant: dict,
        product: dict,
        image_url: str,
        source: str,
        source_id: int | None,
        result: ImageEnrichmentResult,
    ) -> None:
        """Record an image assignment."""
        result.assignments.append(
            {
                "variant_id": variant["id"],
                "product_id": product["id"],
                "product_handle": product["handle"],
                "image_url": image_url,
                "source": source,
                "source_id": source_id,
            }
        )


# ---------------------------------------------------------------------------
# MatrixifyExporter
# ---------------------------------------------------------------------------


class MatrixifyExporter:
    """Export enrichments as Matrixify-compatible CSV for Shopify upload."""

    @staticmethod
    def export_enriched_images(assignments: list[dict]) -> str:
        """Export image enrichment assignments as Matrixify CSV.

        Columns: ID, Handle, Command, Variant ID, Variant Command, Variant Image

        - Command: MERGE on first row per product, blank on subsequent
        - Variant Command: MERGE on all rows
        """
        output = io.StringIO()
        fieldnames = [
            "ID",
            "Handle",
            "Command",
            "Variant ID",
            "Variant Command",
            "Variant Image",
        ]
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()

        # Group by product to set Command correctly
        by_product: dict[int, list[dict]] = {}
        for a in assignments:
            by_product.setdefault(a["product_id"], []).append(a)

        for _product_id, product_assignments in by_product.items():
            for i, a in enumerate(product_assignments):
                writer.writerow(
                    {
                        "ID": a["product_id"],
                        "Handle": a["product_handle"],
                        "Command": "MERGE" if i == 0 else "",
                        "Variant ID": a["variant_id"],
                        "Variant Command": "MERGE",
                        "Variant Image": a["image_url"],
                    }
                )

        return output.getvalue()

    @staticmethod
    def export_descriptions(enrichments: list[dict]) -> str:
        """Export description enrichments as Matrixify CSV.

        Columns: ID, Handle, Command, Body HTML
        """
        output = io.StringIO()
        fieldnames = ["ID", "Handle", "Command", "Body HTML"]
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()

        for e in enrichments:
            writer.writerow(
                {
                    "ID": e["product_id"],
                    "Handle": e["product_handle"],
                    "Command": "MERGE",
                    "Body HTML": e["body_html"],
                }
            )

        return output.getvalue()


# ---------------------------------------------------------------------------
# Variant Sort Order Assessment
# ---------------------------------------------------------------------------


@dataclass
class SortViolation:
    """A single variant sort violation within a product."""

    product_handle: str
    product_title: str
    expected_order: list[str]  # What the order should be
    actual_order: list[str]  # What it actually is


@dataclass
class SortAssessmentResult:
    """Result of variant sort order assessment."""

    correctly_sorted: int = 0
    violations: list[SortViolation] = field(default_factory=list)
    no_position_data: int = 0
    total_products: int = 0


def _variant_label(v: dict) -> str:
    """Create a label for a variant dict for sort display."""
    parts = []
    if v["option1_value"]:
        parts.append(v["option1_value"])
    if v["option2_value"]:
        parts.append(v["option2_value"])
    if v["option3_value"]:
        parts.append(v["option3_value"])
    return " / ".join(parts) if parts else v["sku"] or str(v["id"])


def _size_sort_key(size_str: str) -> tuple:
    """Basic size sort key for variant ordering.

    Orders: XS < S < M < L < XL < XXL, with numeric sizes sorted numerically.
    This replaces the tvr.core.sizes.sort_key dependency.
    """
    canonical = {
        "xxs": (0, 0),
        "xs": (0, 1),
        "s": (0, 2),
        "sm": (0, 2),
        "small": (0, 2),
        "s/m": (0, 3),
        "m": (0, 4),
        "med": (0, 4),
        "medium": (0, 4),
        "m/l": (0, 5),
        "l": (0, 6),
        "lg": (0, 6),
        "large": (0, 6),
        "xl": (0, 7),
        "xxl": (0, 8),
        "2xl": (0, 8),
        "xxxl": (0, 9),
        "3xl": (0, 9),
        "one size": (0, -1),
        "os": (0, -1),
    }
    lower = size_str.strip().lower()
    if lower in canonical:
        return canonical[lower]
    # Try numeric
    try:
        return (1, float(lower))
    except ValueError:
        pass
    # Fallback: alphabetical
    return (2, 0, lower)


def assess_variant_sort_order(
    store: LookoutStore, vendor: str | None = None
) -> SortAssessmentResult:
    """Assess variant sort order against Color-first convention.

    Convention: All sizes of one color before next color, sizes in canonical order.
    """
    result = SortAssessmentResult()

    products = store.list_products(vendor=vendor)

    for product in products:
        variants = store.get_variants(product["id"])
        if len(variants) <= 1:
            continue

        result.total_products += 1

        # Check if any variants have position data
        has_positions = any(v["position"] is not None for v in variants)
        if not has_positions:
            result.no_position_data += 1
            continue

        # Check if product has Color option
        has_color = any(
            v["option1_name"] and v["option1_name"].lower() == "color" for v in variants
        )
        if not has_color:
            # Non-color products: check size ordering
            sorted_by_pos = sorted(variants, key=lambda v: v["position"] or 0)
            actual_sizes = [_variant_label(v) for v in sorted_by_pos]
            expected_sizes = sorted(actual_sizes, key=lambda s: _size_sort_key(s))

            if actual_sizes == expected_sizes:
                result.correctly_sorted += 1
            else:
                result.violations.append(
                    SortViolation(
                        product_handle=product["handle"],
                        product_title=product["title"],
                        expected_order=expected_sizes,
                        actual_order=actual_sizes,
                    )
                )
            continue

        # Color-first ordering: group by color, sort colors alphabetically,
        # within each color sort sizes canonically
        sorted_by_pos = sorted(variants, key=lambda v: v["position"] or 0)
        actual_order = [_variant_label(v) for v in sorted_by_pos]

        # Build expected order: Color-first, sizes within each color
        color_groups: dict[str, list[dict]] = {}
        for v in variants:
            color = v["option1_value"] or "Unknown"
            color_groups.setdefault(color, []).append(v)

        # Determine color order from current positions (first appearance)
        color_order = []
        seen_colors = set()
        for v in sorted_by_pos:
            color = v["option1_value"] or "Unknown"
            if color not in seen_colors:
                color_order.append(color)
                seen_colors.add(color)

        # Build expected: within each color block, sizes should be in canonical order
        expected_order = []
        for color in color_order:
            group = color_groups[color]
            sorted_group = sorted(
                group,
                key=lambda v: _size_sort_key(v["option2_value"] or v["option1_value"] or ""),
            )
            expected_order.extend([_variant_label(v) for v in sorted_group])

        if actual_order == expected_order:
            result.correctly_sorted += 1
        else:
            result.violations.append(
                SortViolation(
                    product_handle=product["handle"],
                    product_title=product["title"],
                    expected_order=expected_order,
                    actual_order=actual_order,
                )
            )

    return result
